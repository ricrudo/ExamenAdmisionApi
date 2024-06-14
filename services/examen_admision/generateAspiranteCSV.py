'''
A partir de excel que envia admisiones, crea un archivo en formato CSV para alimentar al sistema
'''

from pathlib import Path
from openpyxl import load_workbook

grupos = {"GUITARRA ELECTRICA":"GRUPO_1",
          "BAJO ELÉCTRICO":"GRUPO_1",
          "CONTRABAJO":"GRUPO_1",
          "VIOLA":"GRUPO_2",
          "VIOLÍN":"GRUPO_2",
          "VIOLÍN":"GRUPO_2",
          "GUITARRA ACUSTICA":"GRUPO_3",
          "PERCUSIÓN":"GRUPO_4",
          "PIANO":"GRUPO_5",
          "EUFONIO (BOMBARDINO)":"GRUPO_6",
          "TROMPETA":"GRUPO_6",
          "FLAUTA TRAVERSA":"GRUPO_6",
          "TUBA":"GRUPO_6",
          "CANTO":"GRUPO_7",
          }

def getFile():
    while True:
        filename = input("indique el nombre del archivo")
        convocadosPath = Path.cwd() /'data' / filename
        if not convocadosPath.exists():
            print(f"el archivo {convocadosPath.as_posix()} no existe")
            continue
        if convocadosPath.suffix not in ['.xls','.xlsx']:
            print("la extension del archivo no es valida")
            continue
        return load_workbook(convocadosPath.as_posix())

def getSheet(wb):
    if len(wb.sheetnames) == 1:
        return wb.sheetnames[0]
    while True:
        for i, sheet in enumerate(wb.sheetnames, 1):
            print(f"{i} - {sheet}") 
        response = input("Indique la hoja de calculo que contiene la informacion: ")
        try:
            response = int(response)
        except ValueError:
            print("valor no válido")
            continue
        if 0 < response <= len(wb.sheetnames):
            return wb[wb.sheetnames[response-1]]
        print("valor no válido")

def grupoOptions(instrument):
    options = [(v, k) for k, v in grupos.items()]
    while True:
        for i, opt in enumerate(options, 1):
            print(f'{i} - {opt}')
        response = input(f"indique el grupo para el instrumento {instrument}: ")
        try:
            response = int(response)
        except ValueError:
            print("Valor no valido")
            continue
        if 0 < response <= len(options):
            return options[response][0]
        print("Valor no valido")


def getDataPersons(sheet):
    persons = {}
    for row in sheet.rows:
        for cell in row:
            if cell.row < 2:
                break
            if cell.column == 3:
                cedula = str(cell.value).strip()
                persons[cedula] = {}
            elif cell.column == 5:
                persons[cedula]["nombre"] = cell.value.strip()
            elif cell.column == 11:
                instrument = cell.value.strip()
                persons[cedula]["instrumento"] = instrument
                if not instrument:
                    persons[cedula]["grupo"] = ''
                else:
                    grupo = grupos.get(cell.value, None)
                    if not grupo:
                        grupo = grupoOptions(cell.value)
                    persons[cedula]["grupo"] = grupo
            elif cell.column == 12 and not persons[cedula]["grupo"]:
                instrument = cell.value.strip()
                if instrument:
                    persons[cedula]["instrumento"] = instrument
                    grupo = grupos.get(cell.value, None)
                    if not grupo:
                        grupo = grupoOptions(cell.value)
                    persons[cedula]["grupo"] = grupo
    return persons

def csvFormat(persons):
    content = ''
    for person, value in persons.items():
        content += person + ','
        content += value['grupo'] + ','
        content += value['nombre'] + ','
        content += value['instrumento'] + '\n'
    return content[:-1]

def saveFile(csvContent):
    filename = input("Indique el nombre para el archivo a generar: ")
    filePath = Path.cwd() / 'data' / f'{filename}.csv'
    filePath.write_text(csvContent)
    print(f'el archivo {filePath} ha sido guardado con exito')


if __name__ == "__main__":
    wb = getFile()
    sheet = getSheet(wb)
    persons = getDataPersons(sheet)
    csvContent = csvFormat(persons)
    saveFile(csvContent)


