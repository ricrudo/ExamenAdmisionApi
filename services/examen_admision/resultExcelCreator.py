from openpyxl import load_workbook
from pathlib import Path
import requests
import json

def getPlantilla():
	filename = Path.cwd() / 'data' / 'plantilla_resultados.xlsx'
	workbook = load_workbook(filename=filename.as_posix())
	return workbook

def getDatosBase():
	filename = Path.cwd() / 'data' / 'Convocados_Musica_2024_2.xlsx'
	workbook = load_workbook(filename=filename.as_posix())
	return workbook['160 MUSICA']

def casillas(casillaName, row):
	# el primer valor es la columna en la plantilla y el segundo la columna en los datos base
	# si el segundo valor es None, es porque ese dato se debe tomar de los resultados
	casillas = {"DOCUMENTO": ('A', 'C'),
			"NOMBRE": ('B','E'),
			"INSTRUMENTO":('C', None),
			"PRIMERA OPCION": ('D', 'G'),
			"CIRC. REAL": ('H','J'),
			"TEORIA": ('J',None),
			"AUDIOP": ('K',None),
			"LECT. RITM.": ('L',None),
			"LECT. MEL.": ('M',None),
			"OBRA LIBRE": ('N',None),
			"OBRA PREP.": ('O',None),
			"OBRA PRIMERA": ('P',None),
			"LECT. CRIT.": ('Y','R'),
			"MATEMATICAS": ('Z','S'),
			"SOCIALES": ('AA','T'),
			"NATURALES": ('AB','U'),
			"INGLES": ('AC','V'),
			"ICFES": ('AF','Q')}
	plantillaCell = f'{casillas[casillaName][0]}{row}'
	if casillas[casillaName][1]:
		baseCell = f'{casillas[casillaName][1]}{row}'
	else:
		baseCell = None
	return plantillaCell, baseCell

def pasarDatosBase(datosBase, plantilla):
	for cell in datosBase.rows:
		if cell[0].row < 2:
			continue
		if cell[0].value:
			for concepto in ["DOCUMENTO","NOMBRE","PRIMERA OPCION","CIRC. REAL","LECT. CRIT.","MATEMATICAS","SOCIALES","NATURALES","INGLES","ICFES"]:
				cellConcepto = casillas(concepto, cell[0].row)
				plantilla[cellConcepto[0]].value = datosBase[cellConcepto[1]].value

def getResults():
	response = requests.get("https://www.riedmusicapp.com/examination/examen_ua/get_results")
	return response.json()

def agregarGrades(plantilla, results):
	for cell in plantilla.rows:
		if cell[0].row < 2:
			continue
		if cell[0].value:
			currentRow = cell[0].row
			cedula = str(cell[0].value)
			if results[cedula]:
				if results[cedula]['grades_instrumento']:
					libre = []
					preparada = []
					lectura = []
					grades_instrumento = json.loads(results[cedula]['grades_instrumento'])
					for values in grades_instrumento.values():
						if isinstance(values, dict):
							libre.append(values["libre y escalas"])
							preparada.append(values["preparada"])
							lectura.append(values["lectura"])
					if libre:
						cellObraLibre = casillas("OBRA LIBRE", currentRow)[0]
						plantilla[cellObraLibre].value = sum(libre) / len(libre)
					if preparada:
						cellPreparada = casillas("OBRA PREP.", currentRow)[0]
						plantilla[cellPreparada].value = sum(preparada) / len(preparada)
					if lectura:
						cellPrimera = casillas("OBRA PRIMERA", currentRow)[0]
						plantilla[cellPrimera].value = sum(lectura) / len(lectura)
				if results[cedula]['grades_solfeo']:
					grades_solfeo = json.loads(results[cedula]['grades_solfeo'])
					lectura_ritmica = grades_solfeo.get('lectura_ritmica', None)
					if lectura_ritmica is not None:
						cellLecturaRitmica = casillas("LECT. RITM.", currentRow)[0]
						plantilla[cellLecturaRitmica].value = float(lectura_ritmica)
					lectura_melodica = grades_solfeo.get('lectura_melodica', None)
					if lectura_melodica is not None:
						cellLecturaMelodica = casillas("LECT. MEL.", currentRow)[0]
						plantilla[cellLecturaMelodica].value = float(lectura_melodica)
				audioperceptiva = results[cedula].get("audioperceptiva", None)
				if audioperceptiva:
					cellAudioP = casillas("AUDIOP", currentRow)[0]
					plantilla[cellAudioP].value = audioperceptiva
				teoria = results[cedula].get("teoria", None)
				if teoria:
					cellTeoria = casillas("TEORIA", currentRow)[0]
					plantilla[cellTeoria].value = teoria
				response = requests.get(f"https://www.riedmusicapp.com/examination/admisionesUA/services/get_candidate_{cedula}")
				instrument = response.json()[cedula]['group']
				if instrument == "GRUPO_1":
					instrument = "Guitarra"
				elif instrument == "GRUPO_2":
					instrument = "Violin"
				elif instrument == "GRUPO_4":
					instrument = "Percusion"
				elif instrument == "GRUPO_5":
					instrument = "Piano"
				elif instrument == "GRUPO_6":
					instrument = "Viento"
				elif instrument == "GRUPO_7":
					instrument = "Canto"
				else:
					instrument = "Otro"
				cellInstrumento = casillas("INSTRUMENTO", currentRow)[0]
				plantilla[cellInstrumento].value = instrument
				print('done')
				
					
				
	

if __name__ == "__main__":
	results = getResults()
	plantilla = getPlantilla()
	datosBase = getDatosBase()
	pasarDatosBase(datosBase, plantilla['160'])
	agregarGrades(plantilla['160'], results)
	plantilla.save('resultadoFinal.xlsx')
